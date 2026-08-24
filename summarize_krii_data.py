import json
import re
from html.parser import HTMLParser
from pathlib import Path

import openpyxl

base = Path(r"C:\Users\LENOVO\Documents\startup")
bp_path = base / "krii_bp.xlsx"
budget_path = base / "Budget_ProtoMarket_II_AR.xlsx"
bmc_path = base / "krii_bmc_clean.html"


def cell(ws, ref):
    value = ws[ref].value
    return value if value is not None else ""


def nonempty_rows(ws, min_row=1, max_row=None, max_col=None):
    out = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
        vals = [c.value for c in row]
        if any(v not in (None, "") for v in vals):
            out.append(vals)
    return out


class TextParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []

    def handle_data(self, data):
        text = " ".join(data.split())
        if text and not text.startswith(":root") and text != "lucide.createIcons();":
            self.parts.append(text)


parser = TextParser()
parser.feed(bmc_path.read_text(encoding="utf-8", errors="ignore"))
parts = parser.parts
sections = {}
current = None
section_names = {
    "CUSTOMER SEGMENTS",
    "VALUE PROPOSITIONS",
    "CHANNELS",
    "CUSTOMER RELATIONSHIPS",
    "KEY ACTIVITIES",
    "KEY RESOURCES",
    "KEY PARTNERS",
    "COST STRUCTURE",
    "REVENUE STREAMS",
}
for part in parts:
    if part in {"KRII BMC", "BUSINESS MODEL CANVAS"}:
        continue
    if part in section_names:
        current = part
        sections[current] = []
    elif current:
        sections[current].append(part)

bp_formula = openpyxl.load_workbook(bp_path, data_only=False)
bp_values = openpyxl.load_workbook(bp_path, data_only=True)

revenue_ws = bp_formula["A.2. Chiffre dAffaires "]
salary_ws = bp_formula["A.4. Masse Salariale"]
invest_ws = bp_formula["A.1. Investissement"]
pl_ws = bp_values["B.1. P&L"]
cash_ws = bp_values["B.2. TFT"]
fund_ws = bp_values["C. Synthèse Financement "]

revenue_lines = []
for row in [13, 22, 31, 40, 49]:
    name = cell(revenue_ws, f"E{row}")
    if name:
        revenue_lines.append(
            {
                "name": name,
                "qty_y1_to_y5": [cell(revenue_ws, f"{col}{row+2}") for col in "TUVWX"],
                "price_y1_to_y5": [cell(revenue_ws, f"{col}{row+4}") for col in "TUVWX"],
                "sales_y1_to_y5": [cell(revenue_ws, f"{col}{row+6}") for col in "TUVWX"],
            }
        )

salary_lines = []
for row in range(8, 18):
    role = cell(salary_ws, f"B{row}")
    base_salary = cell(salary_ws, f"C{row}")
    if role:
        salary_lines.append({"role": role, "base_monthly_salary": base_salary})

investments = []
for row in range(8, 38):
    item = cell(invest_ws, f"C{row}")
    function = cell(invest_ws, f"D{row}")
    unit = cell(invest_ws, f"E{row}")
    y1 = cell(invest_ws, f"F{row}")
    if item or function or unit or y1:
        investments.append({"item": item, "function": function, "unit_price": unit, "year_1": y1})

budget_formula = openpyxl.load_workbook(budget_path, data_only=False)
budget_values = openpyxl.load_workbook(budget_path, data_only=True)
print("BUDGET_SHEETS", budget_formula.sheetnames)
for sheet in budget_formula.worksheets:
    print("BUDGET_SHEET", sheet.title)
    for row in nonempty_rows(sheet, max_row=35, max_col=12):
        print(row)

summary = {
    "bmc": sections,
    "revenue_lines": revenue_lines,
    "sales_total_y1_to_y5": [cell(revenue_ws, f"{col}57") for col in "TUVWX"],
    "salary_lines": salary_lines,
    "investments": investments,
    "pl_key_rows_sample": nonempty_rows(pl_ws, max_row=50, max_col=10),
    "cash_key_rows_sample": nonempty_rows(cash_ws, max_row=50, max_col=10),
    "funding_summary_sample": nonempty_rows(fund_ws, max_row=50, max_col=10),
}
print("SUMMARY_JSON_START")
print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
