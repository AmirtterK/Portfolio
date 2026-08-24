import re
import zipfile
from html.parser import HTMLParser
from pathlib import Path

files = {
    "pptx": Path(r"C:\Users\LENOVO\Documents\startup\business_plan_template.pptx"),
    "bp": Path(r"C:\Users\LENOVO\Documents\startup\krii_bp.xlsx"),
    "budget": Path(r"C:\Users\LENOVO\Documents\startup\Budget_ProtoMarket_II_AR.xlsx"),
    "bmc": Path(r"C:\Users\LENOVO\Documents\startup\krii_bmc_clean.html"),
}

for key, path in files.items():
    print(key, path.exists(), path.stat().st_size if path.exists() else None)

with zipfile.ZipFile(files["pptx"]) as deck:
    slides = sorted(
        [
            name
            for name in deck.namelist()
            if re.match(r"ppt/slides/slide\d+\.xml$", name)
        ],
        key=lambda name: int(re.search(r"slide(\d+)", name).group(1)),
    )
    print("SLIDES", len(slides))
    for name in slides:
        xml = deck.read(name).decode("utf-8", errors="ignore")
        texts = re.findall(r"<a:t>(.*?)</a:t>", xml)
        texts = [re.sub(r"\s+", " ", text).strip() for text in texts if text.strip()]
        possible = [
            text
            for text in texts
            if re.search(
                r"\{\{|\}\}|\[|\]|placeholder|xxxx|xxx|lorem|enter|insert|company|name|goes here|Business",
                text,
                re.I,
            )
        ]
        print()
        print(name, "texts", len(texts))
        print(" | ".join(texts[:80]))
        if possible:
            print("POSSIBLE:", possible)

try:
    import openpyxl

    for label in ["bp", "budget"]:
        workbook = openpyxl.load_workbook(files[label], data_only=False)
        print()
        print("WORKBOOK", label, workbook.sheetnames)
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows():
                row_values = []
                for cell in row:
                    if cell.value is not None:
                        row_values.append(f"{cell.coordinate}={cell.value}")
                if row_values:
                    rows.append(" ; ".join(row_values))
            print("SHEET", sheet.title, "rows", len(rows))
            for line in rows[:80]:
                print(line[:500])
except Exception as exc:
    print("XLSXERR", repr(exc))


class Parser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []

    def handle_data(self, data):
        text = " ".join(data.split())
        if text:
            self.parts.append(text)


parser = Parser()
parser.feed(files["bmc"].read_text(encoding="utf-8", errors="ignore"))
print()
print("HTML TEXT")
for part in parser.parts[:300]:
    print(part)
